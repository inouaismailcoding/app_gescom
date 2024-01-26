import csv
import json
import openpyxl as px
from django.http import HttpResponse
from django.shortcuts import render
import tablib
from django.http import HttpResponse
from datetime import datetime,timedelta
from django.core.exceptions import ValidationError
from django.utils.translation import gettext as _
import numpy as np
from django.db.models import Sum
from django.utils import timezone
from .models import Article
from transaction.models import TransactionLine,Transaction

def is_valid_date_format(date_string, date_format='%Y-%m-%d'):
    try:
        # Tente de convertir la chaîne de date en objet datetime
        datetime.strptime(date_string, date_format)
        return True
    except ValueError:
        # Gère l'exception si la conversion échoue
        return False


def calculate_stock_article():
    articles = Article.objects.all()
    stock_data = []
    
    total_quantity_out=[]
    total_quantity_in=[]
    date_begin=""
    date_end=""
    # Si on a pas saisie de date de debut
    if not date_begin:
        first_transaction = Transaction.objects.order_by('date').first()
        if first_transaction:
            date_begin = first_transaction.date
    # Si on a pas saisie de date de fin
    if not date_end:
        last_transaction = Transaction.objects.order_by('date').last()
        if last_transaction:
            date_end = last_transaction.date

    # On recupere les données sur la periode selectionnée
    transactions=TransactionLine.objects.filter(transaction__date__range=(date_begin,date_end))
    for article in articles:
        # Calculez la quantité totale entrée pour un article
        quantity_in = TransactionLine.objects.filter(article=article.id,transaction__mouvement=1,transaction__date__range=(date_begin,date_end)).aggregate(Sum('quantity'))['quantity__sum'] or 0

        # On calcul le stock initial 
        initial_stock=0
        begin=Transaction.objects.order_by('date').first().date
        end=date_begin-timedelta(days=1)
        in_periode=TransactionLine.objects.filter(article=article.id,transaction__mouvement=1,transaction__date__range=(begin,end)).aggregate(Sum('quantity'))['quantity__sum'] or 0
        end_periode=TransactionLine.objects.filter(article=article.id,transaction__mouvement=2,transaction__date__range=(begin,end)).aggregate(Sum('quantity'))['quantity__sum'] or 0
        initial_stock=in_periode-end_periode

        # Calculez la quantité totale sortie pour un article
        quantity_out = TransactionLine.objects.filter(article=article.id,transaction__mouvement=2,transaction__date__range=(date_begin,date_end)).aggregate(Sum('quantity'))['quantity__sum'] or 0

        # Calculez la somme de toute les quantités des articles
        total_quantity_out.append(quantity_out)
        total_quantity_in.append(quantity_in)
        row={}

        # Calculez la quantité disponible pour un article
        row['date_begin']= date_begin.date()
        row['article_id']= article.id
        row['reference_article']= article.article_number
        row['article_name']= article.article
        row["initial_stock"]=initial_stock
        row['quantity_in']= quantity_in
        row['quantity_out']= quantity_out
        row['available_quantity']= quantity_in - quantity_out
        row['date_end']= date_end.date()
        stock_data.append(row)
        
    
    """
    total_out=np.array(total_quantity_out)
    total_out=total_out.sum()
    total_in=np.array(total_quantity_in)
    total_in=total_in.sum()
    """
    return report_object_to_excel_file(stock_data)

def report_object_to_excel_file(data):
    # Créer la réponse HTTP appropriée
    response = HttpResponse(content_type=f'text/excel')
    #response = HttpResponse(content_type=f'text/{output_format}')
    # Créer un objet Excel
    wb = px.Workbook()
    # Sélectionner la feuille active (par défaut, c'est la première feuille)
    sheet = wb.active
    # Ecrire le titre du rapport
    title="Rapport de Transaction du date de debut au date de fin "
    sheet.cell(row=3,column=2,value=title)
    # Écrire les en-têtes
    headers = data[0]
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=5, column=col_num, value=header)

    # Écrire les données
    r=6
    c=1
    for d in data:
        for x in d:
            sheet.cell(row=r, column=c, value=d[x])
            c=c+1
        c=1
        r=r+1
    
    # Générer le fichier Excel
    response['Content-Disposition'] = f'attachment; filename="exemple.xlsx"'

    # Sauvegarder le fichier Excel
    wb.save(response)

    return response


def export_data(request,modele):
    # Récupérer le QuerySet à exporter
    queryset = modele.objects.all()  # Vous pouvez appliquer des filtres ou d'autres opérations sur le QuerySet

    # Créer un DataSet avec tablib
    data = tablib.Dataset()
    
    # Ajouter les en-têtes de colonnes au DataSet
    headers = [field.name for field in modele._meta.get_fields()]
    data.headers = headers
    
    # Ajouter les données du QuerySet au DataSet
    
    for obj in queryset:
        data.append([getattr(obj, field) for field in headers]) # Il èya une erreur a ce niveau

    # Définition du nom de fichier pour le téléchargement
    file_name = "export"

    # Exporter en Excel
    excel_data = data.export('xlsx')
    response_excel = HttpResponse(excel_data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response_excel['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'

    # Exporter en CSV
    csv_data = data.export('csv')
    response_csv = HttpResponse(csv_data, content_type='text/csv')
    response_csv['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'

    # Exporter en JSON
    json_data = data.export('json')
    response_json = HttpResponse(json_data, content_type='application/json')
    response_json['Content-Disposition'] = f'attachment; filename="{file_name}.json"'

    return response_excel #, response_csv, response_json  # Retourner toutes les réponses (Excel, CSV, JSON)

def get_data_table_file(request,model):
    # Nom du fichier de sortie
    table_data = model.objects.all().order_by('-id').values()
    file_name='table_data'
    # Format de sortie demandé (csv, json, excel)
    output_format = request.GET.get('format')

    # Créer la réponse HTTP appropriée
    response = HttpResponse(content_type=f'text/{output_format}')

    if output_format == 'csv':
        # Générer le fichier CSV
        response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'
        writer = csv.writer(response)
        # En-têtes
        writer.writerow(table_data[0].keys())
        # Données
        for row in table_data:
            writer.writerow(row.values())
        print(response)

    elif output_format == 'json':
        # Générer le fichier JSON
        response['Content-Disposition'] = f'attachment; filename="{file_name}.json"'
        json.dump(list(table_data), response)
        print(response)

    elif output_format == 'excel':
        # Générer le fichier Excel
        response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
        workbook = px.Workbook()
        sheet = workbook.active

        # En-têtes
        headers = list(table_data[0].keys())
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

        # Données
        for row_num, row_data in enumerate(table_data, 2):
            for col_num, value in enumerate(row_data.values(), 1):
                sheet.cell(row=row_num, column=col_num, value=value)

        workbook.save(response)
        print(response)

    return response




def genererRef(model,d):
    
    try:
        ref=model.objects.all().order_by('id').last().id
        ref+=1
        numero=str(ref).zfill(5)
        numero=d+numero
    except :
        ref=1
        numero=str(ref).zfill(5)
        numero=d+numero
    return numero

def saveRow(texte):
    tab=[]
    row=[]
    if len(texte) > 1:
        for t in texte:
            data2={}
            for m in t.split('&'):
                myrow=m.split('=')
                data2[myrow[0]]=myrow[1]
            row.append(data2)
    else:
        for t in texte:
            data2={}
            for m in t.split('&'):
                myrow=m.split('=')
                data2[myrow[0]]=myrow[1]
            row.append(data2)
    return row

    